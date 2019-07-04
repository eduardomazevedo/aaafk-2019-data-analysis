#!/usr/bin/env python

input1 = './raw-files/apd/Extra_Info_APD_Annotated.xlsx'
input2 = './raw-files/apd/Extra_Info_APD_Antigens-with correct-titles.xlsx'
input3 = './raw-files/apd/APD_centers.csv'
#output = './intermediate-data/APDHistoricalDataAll-sans-MP-cPRA.csv'
input4 = './raw-files/apd/donor_ages.csv'
input5 = './raw-files/apd/recipient_ages.csv'
input6 = './raw-files/apd/donor_sex.csv'
input7 = './raw-files/apd/recipient_sex.csv'

import re, csv
from collections import defaultdict
from itertools   import count
from openpyxl    import load_workbook

# Get filenames from load-globals.do.  This makes the Stata and the python codebases look to the 
# same place for filenames that are shared between them.
with open("./intermediate-data/globals_for_py_code.csv") as globalfile:
    reader = csv.DictReader(globalfile)
    globs  = dict()
    for row in list(reader):
        globs[row['name']]=row['content']

# output  = globs['apd_file_sans_MP_cPRA']
output  = globs['apd_file_pre_conv']

# blood_map = {'O': 3, 'A': 2, 'B': 1, 'AB': 0}
# with open(output, 'w', newline='', encoding='utf_8') as f:
antigens         = ('A1',      'A2',     'B1',      'B2',      'DR1',      'DR2',      'Bw1',      
                    'Bw2',     'Bw4',    'Bw6',     'Cw1',     'Cw2',      'DRW1',     'DRw2', 
                    'DPA1',    'DPA2',   'DPB1',    'DPB2',    'DQA1',     'DQA2',     'DQB1',    
                    'DQB2' )
antigens2        = ('DR51',    'DR52',   'DR53')
antigensTitles   = ('A1',      'A2',     'B1',      'B2',      'DR1',      'DR2',      'Bw1',      
                    'Bw2',     'Bw4',    'Bw6',     'Cw1',     'Cw2',      'DRw1',     'DRw2',
                    'DPa1',    'DPa2',   'DPb1',    'DPb2',    'DQa1',     'DQa2',     'DQb1',
                    'DQb2')
#'a1', 'a2', 'b1', 'b2', 'dr1', 'dr2', 'bw1', 'bw2', 'bw4', 'bw6', 'cw1', 'cw2', 'drw1', 'drw2', 'dpa1', 
#'dpa2', 'dpb1','dpb2', 'dqa1', 'dqa2', 'dqb1', 'dqb2')
antigens2Titles  = ('dr51',    'dr52',   'dr53')
antibodiesTitles = ('antiA',   'antiB',  'antiC',   'antiDR',  'antiDR51', 'antiDR52', 'antiDR53', 
                    'antiBw',  'antiCw', 'antiDRw', 'antiBw4', 'antiBw6',  'antiDPa',  'antiDPb',  
                    'antiDQa', 'antiDQb')
#antibodiesTitles = ('antia', 'antib', 'antic', 'antidr', 'antidr51', 'antidr52', 'antidr53', 
#'antibw', 'anticw', 'antidrw', 'antibw4', 'antibw6', 'antidpa', 'antidpb', 'antidqa', 'antidqb')
antibodies       = ('antiA',   'antiB',  'antiC',   'antiDR',  'antiDR51', 'antiDR52', 'antiDR53', 
                    'antiBw',  'antiCw', 'antiDRw', 'antiBw4', 'antiBw6',  'antiDPA',  'antiDPB', 
                    'antiDQA', 'antiDQB')
rowTitles        = ('id',                'idX',            'famID',         'famIDX',   'chip',  
                    'arr_date_min',      'listingdate',    'arr_date_max',  'age',      'sex',    
                    'race',              'weight',         'type',          'isdonor',  'alias',  
                    'bloodType',         'alt',            'centerID',      'center',   'minWeight', 
                    'donorRelation',     'maxAgeForDonor', 'minHLApoints') +   \
                     antigensTitles +     antigens2Titles + antibodiesTitles + \
                   ('pra',               'cpra',           'donorASubtype', 'Recipient_Non_A1', 
                    'transplanteddate',  'transplanted',   'badData',       'missingPatient', 
                    'dep_date_max',      'dep_date_min',   'tx_id',         'extended_id')
rowTitles       = [ t.lower() for t in rowTitles ]
csvHeaders      =  ('id',                'idX',            'famid',         'famIDX', 
                    'chip',              'arr_date_min',   'listingdate',   'arr_date_max', 
                    'age',               'sex',            'race',          'weight', 
                    'type',              'isdonor',        'alias',         'bloodType', 
                    'alt',               'centerID',       'center',        'minweight', 
                    'donorrelation',     'maxagefordonor', 'minhlapoints') + \
                     antigens +           antigens2 +       antibodies +     \
                    ('pra',              'cpra',           'donorasubtype', 'recipient_non_a1', 
                     'transplanteddate', 'transplanted',   'baddata',       'missingpatient', 
                     'dep_date_max',     'dep_date_min',   'tx_id',         'extended_id')
csvHeaders      = [ h.lower() for h in csvHeaders ]

# iterate rows of input file and for each row yield two dictionaries: donor values and recipient values
# antigens = True to use input2 file, False to use input1 file
def ws_iter(antigens):
    re_ws = re.compile(r'\s+').sub
    ws = load_workbook(input2 if antigens else input1).worksheets[0]
    header1 = header = []  # donor column names
    header2 = []  # recipient column names
    for c in count(1):
        v = ws.cell(row=1, column=c).value  # scan all cells in the first row until an empty one is found
        if not v:
            break
        v = re_ws(' ', v)  # clean whitespace in column names
        if v in ('RECIPIENT ID', 'RECIP ID'):  # at this column recipient data starts
            header = header2
        header.append(v)
    recip_start = len(header1) + 1  # column number at which recipient data starts

    for r in count(2):  # iterate all rows until an empty cell in the first column is found (end of data)
        if not ws.cell(row=r, column=1).value:
            return
        data = []
        for enum in (enumerate(header1, 1),
                     enumerate(header2, recip_start)):  # run separately through donor columns and recipient columns
            if antigens:
                part = {'DRW': set()}  # a set of found DRw51..DRw53 values
                for col, name in enum:
                    v = ws.cell(row=r, column=col).value
                    if name in ('DRB3', 'DRB4', 'DRB5'):
                        # for these columns add their values to ['DRW'] set
                        part['DRW'].add(v)
                    else:
                        # all other column values are put into dictionary
                        part[name] = v
                # copy BW4 column to BW1 and BW6 to BW2
                part['BW1'] = part['BW4']
                part['BW2'] = part['BW6']
            else:
                # create a dictionary of {column name: cell value}
                # python 3.6 dict comprehension
                # part = {name: ws.cell(row=r, column=col).value for col, name in enum}
                part = {}
                for col, name in enum:
                    part[name] = ws.cell(row=r, column=col).value
                # print 12
            data.append(part)
        yield data  # yield a list of two dictionaries (donor's and recipient's)


# regexp finds antigen's name (starting letters) and number (the only number or the one between * and :)
# python 3.6 old patern
# re_digits = re.compile('([^\d*]+)(?:\d*\*)?(\d+)(?::\d+)?').fullmatch
# python 2.7 new patern
def re_digits(string, flags=0):
    return re.match("(?:" + '([^\d*]+)(?:\d*\*)?(\d+)(?::\d+)?' + r")\Z", string, flags=flags)

# update row with antigens from data
def upd_antigens(row, data):
    for k in antigens:
        s = data.get(k)
        if s and s != 'NULL':
            m = re_digits(s.strip())
            #print('debug: {}'.format(m.group(2)))
            # extract only number from this antigen value
            row[k] = str(int(m.group(2)))
        else:  # empty and NULL values get -1
            row[k] = -1
    for k in antigens2:
        # DR51 column gets 1 if DRw51 is present in 'DRW' set. same for DR52, DR53
        row[k] = 1 if k[:2] + 'w' + k[2:] in data['DRW'] else -1


# processing common for donor and recipient
def upd_row(row, data):
    # copy REGISTRATION DATE to RegdateMin, listingDate, RegdateMax
    v = data['REGISTRATION DATE']
    if v:
        v = v.strftime('%m/%d/%Y')
        for k in ('arr_date_min', 'listingdate', 'arr_date_max'):
            row[k] = v
    # copy WITHDRAWAL DATE to departureDate, departureMax
    v = data['WITHDRAWAL DATE']
    if v:
        v = v.strftime('%m/%d/%Y')
        for k in ('dep_date_min', 'dep_date_max'):
            row[k] = v
    # copy XPLANT DATE to TransplantedDate and set Transplanted
    v = data['XPLANT DATE']
    if v:
        row['transplanteddate'] = v.strftime('%m/%d/%Y')
        row['transplanted'] = 1
    # additional IDs: remove first character from original IDs
    row['id'] = row['idX'][1:]
    row['famid'] = row.get('famIDX', '')[1:]


def main():
    with open(output, 'wb') as f:
        writer = csv.DictWriter(f, fieldnames=csvHeaders)
        writerCSV = csv.writer(f)
        writerCSV.writerow(rowTitles)

        #with open(input3) as apdF:
        #    reader = csv.reader(apdF)
    #        for row in reader:
    #            if(row[0].)
#        df = pd.read_csv(input3, names=['ID', 'NAME', 'ABBREVIATION', 'TYPE', 'STATUS'])
#        df2 = df.set_index("ID")

        # Create a center dictionary
        centerDict = dict()
        with open(input3) as apd_centers:
            reader = csv.DictReader(apd_centers)
            for row in reader:
                centerDict['C' + row['ID']] = row['ABBREVIATION'] + '-TX'

        donorAgesDict = dict()
        with open(input4) as donorAges:
            reader = csv.DictReader(donorAges)
            for row in reader:
                donorAgesDict[row['ID']] = row['AGE']

        patientAgesDict = dict()
        with open(input5) as patientAges:
            reader = csv.DictReader(patientAges)
            for row in reader:
                patientAgesDict[row['ID']] = row['AGE']

        donorSexDict = dict()
        with open(input6) as donorSexes:
            reader = csv.DictReader(donorSexes)
            for row in reader:
                donorSexDict[row['ID']] = row['SEX']

        patientSexDict = dict()
        with open(input7) as patientSexes:
            reader = csv.DictReader(patientSexes)
            for row in reader:
                patientSexDict[row['ID']] = row['SEX']

        #writer.writeheader()
        # iterate both input files in parallel, d_data1 is donor's data from input1, r_data1 is recipient's data from input1 etc.
        for (d_data1, r_data1), (d_data2, r_data2) in zip(ws_iter(False), ws_iter(True)):
            ndd = d_data1['PAIR TYPE'] == 'Non-directed'
            if not ndd:
                # recipient row is output if donor is not ndd
                row = {
                    'idX': r_data1['RECIP ID'],
                    'extended_id': r_data1['RECIP ID'],
                    'famIDX': r_data1['RECIP ID'],
                    'isdonor': 0,
                    'centerID': r_data1['RECIP CENTER'],
                    'bloodType': r_data1['RECIP ABO'],
                    'tx_id': r_data1['XPLANT DONOR'],
                    'cpra': r_data1['PRA'],
                }
                upd_antigens(row, r_data2)
                s = r_data2['RECIPIENT UNACCEPTABLE ANTIGENS']
                if s and s != 'NULL':
                    ab = defaultdict(list)
                    # split antibodies by pipe
                    for x in s.split('|'):
                        if x:
                            x = x.strip()
                            m = re_digits(x)
                            name = m.group(1)  # extract antibody name
                            if name == 'DRB':  # treat DRB as DR
                                name = 'DR'
                            # append antibody number to a list under antibody name in ab dictionary
                            ab[name].append(str(int(m.group(2))))
                            # DRw51 atibody additionaly produces 1 in antiDR51 column, same for DRw52, DRw53
                            if x in ('DRw51', 'DRw52', 'DRw53'):
                                row['antiDR' + x[-2:]] = 1
                    # join antibody numbers with pipe and put the value in column with antibody name
                    for k, v in ab.items():
                        row['anti' + k] = '|'.join(v)

                upd_row(row, r_data1)
                centerShort = row['centerID'][1:]
                #print centerShort
                row['center'] = centerDict[row['centerID']]
                row['age'] = patientAgesDict[row['id']]
                row['sex'] = patientSexDict[row['id']]
                writer.writerow( dict( (k.lower(), row[k]) for k in row.keys() ) )

            # donor row is always output
            row = {
                'idX': d_data1['DONOR ID'],
                'extended_id': d_data1['DONOR ID'],
                'isdonor': 1,
                'centerID': d_data1['DONOR CENTER'],
                'bloodType': d_data1['DONOR ABO'],
                'tx_id': d_data1['XPLANT RECIP'],
            }
            if not ndd:
                row['famIDX'] = r_data1['RECIP ID']
            else:
                row['alt'] = 1
            centerShort = row['centerID'][1:]
            #print centerShort

            row['center'] = centerDict[row['centerID']]

            # 'ABBREVIATION']
            upd_antigens(row, d_data2)
            upd_row(row, d_data1)
            row['age'] = donorAgesDict[row['id']]
            row['sex'] = donorSexDict[row['id']]
            writer.writerow( dict( (k.lower(), row[k]) for k in row.keys() ) )


if __name__ == '__main__':
    main()
